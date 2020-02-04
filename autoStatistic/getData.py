import openpyxl
import re
import os

def getDayItems(dataPath):
    wb = openpyxl.load_workbook(dataPath)
    ws = wb['Sheet1']
    items = {}

    pattern = re.compile(r'.*\\(.+)_(.+)\.xlsx')
    aa = pattern.search(dataPath)
    items['月'] = aa[1]
    items['日'] = aa[2]

    for n in range(18):
        if (ws.cell(3+n,2).value != 0) and (ws.cell(3+n,2).value != None):
            items[ws.cell(3+n,1).value] = ws.cell(3+n,2).value
    for n in range(18):
        if (ws.cell(3+n,4).value != 0) and (ws.cell(3+n,4).value != None):
            items[ws.cell(3+n,3).value] = ws.cell(3+n,4).value
    for n in range(18):
        if (ws.cell(3+n,6).value != 0) and (ws.cell(3+n,6).value != None):
            items[ws.cell(3+n,5).value] = ws.cell(3+n,6).value
    for n in range(15):
        if (ws.cell(3+n,8).value != 0) and (ws.cell(3+n,8).value != None):
            items[ws.cell(3+n,7).value] = ws.cell(3+n,8).value
    for n in range(12):
        if (ws.cell(23+n,2).value != 0) and (ws.cell(23+n,2).value != None):
            items[ws.cell(23+n,1).value] = ws.cell(23+n,2).value

    return items

def getAllItems(rootPath, allItems=[]):
    files = os.listdir(rootPath)
    for file in files:
        dataPath = rootPath + '\\' + file
        items = getDayItems(dataPath)
        allItems.append(items)
    return allItems

def checkRule(ws, data):
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            if ws.cell(row+1,col+1).value == data:
                return True
    return False

def checkItems(rulePath, allItems=[]):
    wb = openpyxl.load_workbook(rulePath)
    ws = wb['Sheet1']
    checkResult = True
    
    for n in range(len(allItems)):
        for item in allItems[n].keys():
            if checkRule(ws, item) == False:
                if (item != '月') and (item != '日'):
                    print(allItems[n]['月']+'月'+allItems[n]['日']+'日的【'+item+'】不在规则列表中。')
                    checkResult = False
    
    return checkResult

def fillTable(tablePath, allItems=[]):
    pass

if __name__ == "__main__":
    allItems = []
    allItems = getAllItems(r'./xx医院/单子', allItems)
    checkResult = checkItems(r'./xx医院/规则.xlsx',allItems)
    if checkResult == True:
        pass
    

